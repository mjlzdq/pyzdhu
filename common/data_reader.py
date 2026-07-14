"""
数据驱动测试 - 文件读取器
支持 CSV (.csv) 和 Excel (.xlsx) 格式
"""
import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Optional


class DataReader:
    """读取 xlsx / csv 测试数据文件，自动识别格式"""

    @staticmethod
    def read_csv(filepath: str, encoding: str = "utf-8-sig", auto_convert: bool = True) -> List[Dict[str, Any]]:
        """
        读取 CSV 文件，每行返回一个 dict

        自动处理：
        - BOM 头 (utf-8-sig)
        - JSON 字段自动解析（以 { 或 [ 开头的值）
        - 数字自动转换

        Args:
            auto_convert: 是否自动转换数据类型（默认 True）；设为 False 时所有值保留字符串
        """
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"数据文件不存在: {filepath}")

        rows = []
        with open(filepath, "r", encoding=encoding, newline="") as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, start=2):
                # 跳过空行
                if all(v.strip() == "" for v in row.values()):
                    continue
                parsed = {}
                for key, value in row.items():
                    key = key.strip()
                    value = value.strip() if value else ""
                    parsed[key] = DataReader._auto_convert(value, auto_convert=auto_convert)
                parsed["_row_num"] = row_num  # 记录行号，方便定位
                rows.append(parsed)

        if not rows:
            raise ValueError(f"数据文件为空或格式不正确: {filepath}")

        return rows

    @staticmethod
    def read_xlsx(filepath: str, sheet: Optional[str] = None, auto_convert: bool = True) -> List[Dict[str, Any]]:
        """
        读取 Excel 文件，每行返回一个 dict

        Args:
            filepath: xlsx 文件路径
            sheet: 工作表名，默认第一个 sheet
            auto_convert: 是否自动转换数据类型（默认 True）；设为 False 时所有值保留字符串
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("读取 xlsx 需要 openpyxl，请执行: pip install openpyxl")

        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"数据文件不存在: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet] if sheet else wb.active

        # 读取表头（第一行）
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(1, col).value
            headers.append(str(cell_value).strip() if cell_value else f"col_{col}")

        # 读取数据行
        rows = []
        for row_num in range(2, ws.max_row + 1):
            row_data = {}
            is_empty = True
            for col, header in enumerate(headers, start=1):
                cell_value = ws.cell(row_num, col).value
                if cell_value is not None:
                    is_empty = False
                value = str(cell_value).strip() if cell_value is not None else ""
                row_data[header] = DataReader._auto_convert(value, auto_convert=auto_convert)

            if not is_empty:
                row_data["_row_num"] = row_num
                rows.append(row_data)

        wb.close()

        if not rows:
            raise ValueError(f"数据文件为空或格式不正确: {filepath}")

        return rows

    @staticmethod
    def read(filepath: str, auto_convert: bool = True, **kwargs) -> List[Dict[str, Any]]:
        """
        自动识别文件格式并读取

        支持:
        - .csv  → read_csv
        - .xlsx → read_xlsx

        Args:
            auto_convert: 是否自动转换数据类型（默认 True）；设为 False 时所有值保留字符串
        """
        filepath = Path(filepath)
        suffix = filepath.suffix.lower()

        if suffix == ".csv":
            return DataReader.read_csv(str(filepath), auto_convert=auto_convert, **kwargs)
        elif suffix in (".xlsx", ".xls"):
            return DataReader.read_xlsx(str(filepath), auto_convert=auto_convert, **kwargs)
        else:
            raise ValueError(f"不支持的文件格式: {suffix}，仅支持 .csv / .xlsx")

    @staticmethod
    def _auto_convert(value: str, auto_convert: bool = True) -> Any:
        """自动转换值类型：JSON 字符串 → dict/list，数字 → int/float"""
        if not auto_convert:
            return value
        if not value:
            return ""

        # 尝试解析 JSON
        if value.startswith("{") or value.startswith("["):
            try:
                return json.loads(value)
            except (json.JSONDecodeError, ValueError):
                pass

        # 尝试转换数字
        try:
            if "." in value:
                return float(value)
            return int(value)
        except ValueError:
            pass

        # 布尔值
        if value.lower() in ("true", "false"):
            return value.lower() == "true"

        # None / null
        if value.lower() in ("none", "null"):
            return None

        return value

    @staticmethod
    def detect_format(rows: List[Dict]) -> str:
        """
        自动检测数据文件格式类型

        返回:
        - "simple"  : 简单模式（含 title/body/userId 字段）
        - "generic" : 通用模式（含 method/path/request_body 字段）
        - "unknown" : 无法识别
        """
        if not rows:
            return "unknown"

        headers = set(rows[0].keys())
        # 去掉内部字段
        headers.discard("_row_num")

        # 通用模式检测
        generic_keys = {"method", "path", "expected_status"}
        if generic_keys.issubset(headers):
            return "generic"

        # 简单模式检测
        simple_keys = {"title", "expected_status"}
        if simple_keys.issubset(headers):
            return "simple"

        return "unknown"
