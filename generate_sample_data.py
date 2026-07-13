#!/usr/bin/env python3
"""
生成示例测试数据文件（CSV → XLSX 转换）

用法:
    python3 generate_sample_data.py          # 生成所有示例数据
    python3 generate_sample_data.py --csv    # 只生成 CSV
    python3 generate_sample_data.py --xlsx   # 只生成 XLSX
"""
import csv
import sys
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"


def generate_posts_xlsx():
    """从 sample_posts.csv 生成 sample_posts.xlsx"""
    try:
        import openpyxl
    except ImportError:
        print("⚠️ 需要 openpyxl 来生成 xlsx 文件，正在尝试安装...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    csv_path = DATA_DIR / "sample_posts.csv"
    xlsx_path = DATA_DIR / "sample_posts.xlsx"

    if not csv_path.exists():
        print(f"❌ CSV 文件不存在: {csv_path}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Posts"

    # 写入表头和数据
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row_idx, col_idx, value.strip() if value else "")

    # 调整列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # 冻结首行（表头）
    ws.freeze_panes = "A2"

    # 添加第二个 sheet：通用模式示例
    ws2 = wb.create_sheet("Generic")
    csv_path2 = DATA_DIR / "sample_generic.csv"
    if csv_path2.exists():
        with open(csv_path2, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws2.cell(row_idx, col_idx, value.strip() if value else "")
        ws2.freeze_panes = "A2"
        for col in ws2.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(str(xlsx_path))
    print(f"✅ XLSX 文件已生成: {xlsx_path}")
    print(f"   Sheet1: Posts ({ws.max_row - 1} 条数据)")
    print(f"   Sheet2: Generic ({ws2.max_row - 1} 条数据)")


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    args = sys.argv[1:]
    gen_csv = "--csv" in args or not args
    gen_xlsx = "--xlsx" in args or not args

    if gen_csv:
        print("CSV 文件已存在于 data/ 目录")

    if gen_xlsx:
        generate_posts_xlsx()


if __name__ == "__main__":
    main()
