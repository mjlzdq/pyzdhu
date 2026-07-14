"""
接口批量测试平台 - 后端 FastAPI v3.0
功能：CSV/XLSX 批量测试 | 变量提取与链路 | 动态占位符 | 结果导出 Excel | 单接口调试
"""
import csv
import io
import json
import random
import re
import string
import sys
import time
import uuid as uuid_lib
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, List, Dict, Any

# 将项目根目录加入 sys.path，以便引入 common 模块
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

from common.http_client import (
    HttpClient, HttpError, HttpTimeoutError, HttpMaxRetryError,
)
from common.utils import get_nested

app = FastAPI(title="接口批量测试平台", version="2.0")

# 静态文件（CSS 等零外网资源）
FRONTEND_DIR = Path(__file__).resolve().parent.parent / "frontend"
app.mount("/static", StaticFiles(directory=str(FRONTEND_DIR)), name="static")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

FRONTEND_DIR.mkdir(parents=True, exist_ok=True)


@app.get("/")
async def index():
    html_path = FRONTEND_DIR / "index.html"
    if html_path.exists():
        return HTMLResponse(html_path.read_text(encoding="utf-8"))
    return HTMLResponse("<h1>前端文件未找到</h1>")


# ═══════════════════════════════════════════════════════════════
# 数据解析
# ═══════════════════════════════════════════════════════════════

def parse_csv(content: str) -> List[Dict]:
    """解析 CSV 内容，自动识别 JSON 字段，跳过全空行"""
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for row in reader:
        cleaned = {
            k.strip(): (v.strip() if isinstance(v, str) else v)
            for k, v in row.items()
        }
        # 跳过全空行：所有值都是空字符串或 None
        if all(not v or (isinstance(v, str) and not v.strip()) for v in cleaned.values()):
            continue
        cleaned = _auto_parse_json(cleaned)
        rows.append(cleaned)
    return rows


def parse_xlsx(content: bytes) -> List[Dict]:
    """解析 XLSX 内容，第一行为表头"""
    wb = openpyxl.load_workbook(BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # 表头
    headers = [
        str(h).strip() if h is not None else f"col_{i}"
        for i, h in enumerate(next(rows_iter))
    ]

    rows = []
    for row_values in rows_iter:
        if all(v is None for v in row_values):
            continue
        row = {}
        for i, header in enumerate(headers):
            val = row_values[i] if i < len(row_values) else ""
            row[header] = str(val).strip() if val is not None else ""
        row = _auto_parse_json(row)
        rows.append(row)

    wb.close()
    return rows


def _auto_parse_json(row: Dict) -> Dict:
    """自动将 JSON 字符串字段转为 dict/list"""
    for key in list(row.keys()):
        val = row[key]
        if isinstance(val, str) and (val.startswith("{") or val.startswith("[")):
            try:
                row[key] = json.loads(val)
            except (json.JSONDecodeError, ValueError):
                pass
    return row


def detect_format(filename: str) -> str:
    if filename.lower().endswith('.xlsx'):
        return 'xlsx'
    return 'csv'


# ═══════════════════════════════════════════════════════════════
# 动态占位符引擎
# ═══════════════════════════════════════════════════════════════

def _render_dynamic_placeholders(text: str) -> str:
    """替换内置动态占位符"""
    if not isinstance(text, str):
        return text

    text = text.replace("{{uuid}}", str(uuid_lib.uuid4()))
    text = text.replace("{{uuid_short}}", str(uuid_lib.uuid4())[:8])
    text = text.replace("{{now}}", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    text = text.replace("{{date}}", datetime.now().strftime("%Y-%m-%d"))
    text = text.replace("{{timestamp}}", str(int(time.time())))

    # {{random_int}} 或 {{random_int:1000:9999}}
    for m in re.finditer(r'\{\{random_int(?::(\d+):(\d+))?\}\}', text):
        lo = int(m.group(1)) if m.group(1) else 0
        hi = int(m.group(2)) if m.group(2) else 999999
        text = text.replace(m.group(0), str(random.randint(lo, hi)))

    # {{random_str}} 或 {{random_str:12}}
    for m in re.finditer(r'\{\{random_str(?::(\d+))?\}\}', text):
        length = int(m.group(1)) if m.group(1) else 8
        text = text.replace(
            m.group(0),
            ''.join(random.choices(string.ascii_letters + string.digits, k=length))
        )

    # {{faker_email}}
    if "{{faker_email}}" in text:
        name = ''.join(random.choices(string.ascii_lowercase, k=8))
        text = text.replace("{{faker_email}}", f"{name}@autotest.com")

    # {{faker_name}}
    if "{{faker_name}}" in text:
        first = random.choice("张李王赵陈刘杨黄周吴")
        last = random.choice("伟芳娜敏静强磊洋勇军")
        text = text.replace("{{faker_name}}", f"{first}{last}")

    # {{faker_phone}}
    if "{{faker_phone}}" in text:
        text = text.replace("{{faker_phone}}", f"1{random.randint(3,9)}{random.randint(100000000,999999999)}")

    return text


def _render_all_placeholders(data, pool: Dict[str, Any]) -> Any:
    """递归渲染变量池 {{var}} + 动态占位符"""
    if isinstance(data, str):
        for k, v in pool.items():
            data = data.replace(f"{{{{{k}}}}}", str(v))
        data = _render_dynamic_placeholders(data)
        return data
    elif isinstance(data, dict):
        return {k: _render_all_placeholders(v, pool) for k, v in data.items()}
    elif isinstance(data, list):
        return [_render_all_placeholders(item, pool) for item in data]
    return data


# ═══════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════

def _truncate(data, max_len=2000):
    s = json.dumps(data, ensure_ascii=False) if isinstance(data, (dict, list)) else str(data)
    if len(s) > max_len:
        return s[:max_len] + "...(已截断)"
    return s


# ═══════════════════════════════════════════════════════════════
# 测试运行器
# ═══════════════════════════════════════════════════════════════

class TestRunner:
    """单次批量测试运行器，管理变量池、执行、断言"""

    def __init__(self, base_url: str, cookie: str, timeout: int):
        self.base_url = base_url.rstrip('/')
        self.cookie = cookie
        self.timeout = timeout
        self.pool: Dict[str, Any] = {}

    def run(self, rows: List[Dict]) -> Dict:
        results = []
        status_counts = {"成功": 0, "失败": 0, "异常": 0}
        req_info = {}  # 用于异常时的兜底信息

        for idx, row in enumerate(rows):
            case_id = row.get("case_id", f"行{idx + 1}")
            desc = row.get("description", "")
            start = time.time()
            req_info = {"method": "?", "url": "?"}

            try:
                # 1) 渲染占位符（变量池 + 动态）
                row = _render_all_placeholders(row, self.pool)

                # 2) 构建请求
                req = self._build_request(row)
                req_info = {"method": req["method"], "url": req["url"]}
                headers = req["headers"]
                if self.cookie:
                    headers["Cookie"] = self.cookie

                # 3) 发送
                response = requests.request(
                    method=req["method"],
                    url=req["url"],
                    headers=headers,
                    json=req["json"],
                    timeout=self.timeout,
                )
                elapsed_ms = round((time.time() - start) * 1000)

                # 4) 解析响应
                try:
                    resp_body = response.json()
                except Exception:
                    resp_body = response.text

                # 5) 提取变量
                extract_spec = row.get("extract", "")
                if extract_spec and isinstance(extract_spec, str):
                    self._extract_vars(extract_spec, resp_body)

                # 6) 断言
                passed, failure_reason = self._assert(
                    row, response, resp_body, elapsed_ms
                )

                if passed:
                    status_counts["成功"] += 1
                else:
                    status_counts["失败"] += 1

                results.append({
                    "case_id": case_id,
                    "description": desc,
                    "method": req["method"],
                    "url": req["url"],
                    "status_code": response.status_code,
                    "elapsed_ms": elapsed_ms,
                    "passed": passed,
                    "failure_reason": failure_reason,
                    "response_body": _truncate(resp_body),
                    "response_full": (
                        json.dumps(resp_body, ensure_ascii=False)
                        if isinstance(resp_body, (dict, list))
                        else str(resp_body)
                    ),
                })

            except Exception as e:
                status_counts["异常"] += 1
                elapsed = round((time.time() - start) * 1000)
                results.append({
                    "case_id": case_id,
                    "description": desc,
                    "method": req_info["method"],
                    "url": req_info["url"],
                    "status_code": 0,
                    "elapsed_ms": elapsed,
                    "passed": False,
                    "failure_reason": str(e),
                    "response_body": None,
                    "response_full": str(e),
                })

        return {
            "total": len(rows),
            "status_counts": status_counts,
            "results": results,
            "variable_pool": self.pool,
            "csv_columns": list(rows[0].keys()) if rows else [],
        }

    def _build_request(self, row: Dict) -> Dict:
        method = row.get("method", "GET").upper()
        path = row.get("path", "/")

        if path.startswith("http"):
            url = path
        elif path.startswith("/"):
            url = self.base_url + path
        else:
            url = self.base_url + "/" + path

        body = None
        if "request_body" in row:
            rb = row["request_body"]
            if isinstance(rb, dict):
                body = rb
            elif isinstance(rb, str) and rb.strip():
                try:
                    body = json.loads(rb)
                except json.JSONDecodeError:
                    body = None

        headers = {"Content-Type": "application/json"}
        if "headers" in row:
            h = row["headers"]
            if isinstance(h, dict):
                headers.update(h)
            elif isinstance(h, str) and h.strip():
                try:
                    headers.update(json.loads(h))
                except json.JSONDecodeError:
                    pass

        return {"method": method, "url": url, "headers": headers, "json": body}

    def _extract_vars(self, spec: str, resp_body: Any):
        """
        提取变量到变量池
        格式: token=data.token | order_id=data.list[0].id
        """
        for pair in spec.split("|"):
            pair = pair.strip()
            if "=" not in pair:
                continue
            name, path = pair.split("=", 1)
            name = name.strip()
            path = path.strip()
            val = get_nested(resp_body, path)
            if val is not None:
                self.pool[name] = val

    def _assert(
        self, row: Dict, response, resp_body: Any, elapsed_ms: int
    ) -> tuple:
        failures = []

        # 状态码
        expected_status = row.get("expected_status", "")
        if expected_status:
            try:
                if response.status_code != int(expected_status):
                    failures.append(
                        f"状态码: 期望={expected_status}, 实际={response.status_code}"
                    )
            except ValueError:
                pass

        # 响应时间
        expected_max_ms = row.get("expected_max_ms", "")
        if expected_max_ms:
            try:
                max_ms = int(expected_max_ms)
                if elapsed_ms > max_ms:
                    failures.append(
                        f"响应超时: 期望≤{max_ms}ms, 实际={elapsed_ms}ms"
                    )
            except ValueError:
                pass

        # 字段断言: expected_xxx → 检查 resp_body 中的 xxx 字段
        for key in list(row.keys()):
            if key.startswith("expected_") and key not in (
                "expected_status",
                "expected_max_ms",
            ):
                field_name = key[len("expected_"):]
                expected_val = row[key]
                actual_val = get_nested(resp_body, field_name)

                # 统一转字符串比较
                exp = str(expected_val) if expected_val is not None else ""
                act = str(actual_val) if actual_val is not None else ""
                if exp and exp != act:
                    failures.append(
                        f"[{field_name}] 期望={exp}, 实际={act}"
                    )

        if failures:
            return False, "; ".join(failures)
        return True, ""


# ═══════════════════════════════════════════════════════════════
# API 端点
# ═══════════════════════════════════════════════════════════════

@app.post("/api/run")
async def run_tests(
    file: Optional[UploadFile] = File(None),
    baseUrl: str = Form(""),
    cookie: str = Form(""),
    timeout: int = Form(30),
    rows_json: str = Form(""),
    init_vars_json: str = Form(""),
):
    """
    批量执行接口测试
    支持 CSV / XLSX 上传 或 手动添加用例（rows_json），自动渲染占位符、提取变量、执行断言
    """
    # 优先使用前端编辑数据，其次从文件解析
    if rows_json:
        rows = json.loads(rows_json)
    elif file and file.filename:
        content = await file.read()
        filename = file.filename or ""
        if detect_format(filename) == 'xlsx':
            rows = parse_xlsx(content)
        else:
            rows = parse_csv(content.decode("utf-8"))
    else:
        raise HTTPException(400, "请上传文件或手动添加用例")

    if not rows:
        raise HTTPException(400, "文件为空或格式不正确")

    runner = TestRunner(baseUrl, cookie, timeout)

    # 初始变量（用户手动设置的）
    if init_vars_json:
        try:
            init_vars = json.loads(init_vars_json)
            if isinstance(init_vars, dict):
                runner.pool.update(init_vars)
        except json.JSONDecodeError:
            pass

    result = runner.run(rows)
    return result


@app.post("/api/preview")
async def preview_file(file: UploadFile = File(...)):
    """预览文件列名和前 3 条内容"""
    content = await file.read()
    filename = file.filename or ""

    if detect_format(filename) == 'xlsx':
        rows = parse_xlsx(content)
    else:
        rows = parse_csv(content.decode("utf-8"))

    # 将复杂对象序列化为字符串，方便前端展示
    safe_preview = []
    for r in rows[:3]:
        safe = {}
        for k, v in r.items():
            safe[k] = json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v
        safe_preview.append(safe)

    return {
        "total": len(rows),
        "columns": list(rows[0].keys()) if rows else [],
        "preview": safe_preview,
        "filename": filename,
    }


@app.post("/api/preview_full")
async def preview_full(file: UploadFile = File(...)):
    """返回文件全部行数据（用于前端编辑）"""
    content = await file.read()
    filename = file.filename or ""

    if detect_format(filename) == 'xlsx':
        rows = parse_xlsx(content)
    else:
        rows = parse_csv(content.decode("utf-8"))

    # 将复杂对象序列化
    safe_rows = []
    for r in rows:
        safe = {}
        for k, v in r.items():
            safe[k] = json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v
        safe_rows.append(safe)

    return {
        "total": len(safe_rows),
        "columns": list(safe_rows[0].keys()) if safe_rows else [],
        "rows": safe_rows,
    }


@app.post("/api/export")
async def export_results(results_json: str = Form(...)):
    """
    导出测试结果为 Excel
    包含：汇总概览 + 详细结果 + 变量池
    """
    data = json.loads(results_json)
    results = data.get("results", [])
    summary = data.get("status_counts", {})
    total = data.get("total", 0)
    pool = data.get("variable_pool", {})

    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    pass_fill = PatternFill(start_color="E6FFED", end_color="E6FFED", fill_type="solid")
    fail_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

    # ── Sheet 1: 汇总 ──
    ws1 = wb.active
    ws1.title = "汇总概览"
    ws1.merge_cells('A1:B1')
    ws1['A1'] = "接口批量测试报告"
    ws1['A1'].font = Font(bold=True, size=16, color="667EEA")
    ws1['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    ws1['A4'] = "指标"; ws1['B4'] = "数值"
    for cell in [ws1['A4'], ws1['B4']]:
        cell.fill = header_fill; cell.font = header_font; cell.border = thin_border

    metrics = [
        ("总用例数", total),
        ("成功", summary.get("成功", 0)),
        ("失败", summary.get("失败", 0)),
        ("异常", summary.get("异常", 0)),
    ]
    pass_rate = round(summary.get("成功", 0) / max(total, 1) * 100, 1)
    metrics.append(("通过率", f"{pass_rate}%"))

    for i, (label, value) in enumerate(metrics):
        ws1[f'A{5 + i}'] = label; ws1[f'B{5 + i}'] = value
        ws1[f'A{5 + i}'].border = thin_border
        ws1[f'B{5 + i}'].border = thin_border

    # 变量池
    if pool:
        row_offset = 5 + len(metrics) + 1
        ws1.merge_cells(f'A{row_offset}:B{row_offset}')
        ws1[f'A{row_offset}'] = "变量池"
        ws1[f'A{row_offset}'].font = Font(bold=True, size=12)
        row_offset += 1
        ws1[f'A{row_offset}'] = "变量名"; ws1[f'B{row_offset}'] = "变量值"
        for cell in [ws1[f'A{row_offset}'], ws1[f'B{row_offset}']]:
            cell.fill = header_fill; cell.font = header_font; cell.border = thin_border
        row_offset += 1
        for k, v in pool.items():
            ws1[f'A{row_offset}'] = k; ws1[f'B{row_offset}'] = str(v)
            ws1[f'A{row_offset}'].border = thin_border
            ws1[f'B{row_offset}'].border = thin_border
            row_offset += 1

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 40

    # ── Sheet 2: 详细结果 ──
    ws2 = wb.create_sheet("详细结果")
    detail_headers = [
        "序号", "用例ID", "描述", "方法", "URL", "状态码",
        "耗时(ms)", "结果", "失败原因", "响应体"
    ]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

    for r_idx, r in enumerate(results):
        row_num = r_idx + 2
        values = [
            r_idx + 1,
            r.get("case_id", ""),
            r.get("description", ""),
            r.get("method", ""),
            r.get("url", ""),
            r.get("status_code", ""),
            r.get("elapsed_ms", ""),
            "✓ 通过" if r.get("passed") else "✗ 失败",
            r.get("failure_reason", ""),
            (r.get("response_full", "") or "")[:800],
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws2.cell(row=row_num, column=col_idx, value=v)
            cell.border = thin_border
            if col_idx == 8:
                cell.fill = pass_fill if r.get("passed") else fail_fill

    col_widths = [6, 22, 18, 8, 50, 8, 10, 10, 35, 50]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=test_report.xlsx"
        },
    )


# ═══════════════════════════════════════════════════════════════
# 单接口调试（复用 common.http_client.HttpClient 的重试/超时逻辑）
# ═══════════════════════════════════════════════════════════════

class _AttemptLogSession(requests.Session):
    """包装 requests.Session，记录每次请求尝试，供 Web 平台展示重试过程。"""

    def __init__(self):
        super().__init__()
        self.attempts: list = []

    def request(self, method, url, **kwargs):
        t0 = time.time()
        try:
            resp = super().request(method, url, **kwargs)
            self.attempts.append({
                "attempt": len(self.attempts) + 1,
                "status_code": resp.status_code,
                "elapsed_ms": round((time.time() - t0) * 1000),
                "retried": False,
                "error": None,
            })
            return resp
        except (requests.Timeout, requests.ConnectionError) as e:
            self.attempts.append({
                "attempt": len(self.attempts) + 1,
                "status_code": 0,
                "elapsed_ms": round((time.time() - t0) * 1000),
                "retried": False,
                "error": f"{type(e).__name__}: {e}",
            })
            raise


class _DebugHttpClient(HttpClient):
    """扩展 HttpClient，使用带尝试日志的 Session。"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # 用空日志会话替换默认会话，保证返回的请求头仅含用户传入的头
        self.session = _AttemptLogSession()


def _finalize_attempts(attempts: list) -> None:
    """
    回填 retried 标记。

    语义与原始重试逻辑一致：若某次尝试之后还有后续尝试，
    说明它触发了重试（retried=True），最后一次尝试必然为 False。
    """
    last = len(attempts) - 1
    for i, attempt in enumerate(attempts):
        attempt["retried"] = i < last


def _single_request(
    method: str,
    url: str,
    headers: Optional[Dict[str, str]] = None,
    body: Optional[str] = None,
    timeout: int = 30,
    retry: int = 3,
    retry_delay: float = 1.0,
) -> Dict[str, Any]:
    """
    发送单次 HTTP 请求，带重试机制（复用 common.http_client.HttpClient）。
    返回统一结构，包含请求详情、响应和耗时。
    """
    client = _DebugHttpClient(
        base_url="", timeout=timeout, retry=retry, retry_delay=retry_delay,
    )
    if headers:
        client.session.headers.update(headers)

    req_body_text = body or ""
    json_body = None
    if req_body_text.strip():
        try:
            json_body = json.loads(req_body_text)
        except (json.JSONDecodeError, ValueError):
            pass  # 非 JSON body 保持 None

    start = time.time()
    try:
        resp = client._request(
            method, url,
            json=json_body,
            data=None if json_body else (req_body_text or None),
        )
    except HttpError as e:
        _finalize_attempts(client.session.attempts)
        return {
            "success": False,
            "request": {
                "method": method.upper(),
                "url": url,
                "headers": dict(client.session.headers),
                "body": req_body_text,
            },
            "error": str(e),
            "attempts": client.session.attempts,
            "retry_used": any(a.get("retried") for a in client.session.attempts),
            "elapsed_ms": round((time.time() - start) * 1000),
        }
    finally:
        client.close()

    _finalize_attempts(client.session.attempts)
    try:
        resp_json = resp.json()
    except Exception:
        resp_json = None

    return {
        "success": True,
        "request": {
            "method": method.upper(),
            "url": url,
            "headers": dict(client.session.headers),
            "body": req_body_text,
        },
        "response": {
            "status_code": resp.status_code,
            "headers": dict(resp.headers),
            "body_text": resp.text[:50000],
            "body_json": resp_json,
            "elapsed_ms": round((time.time() - start) * 1000),
            "content_length": len(resp.content),
        },
        "attempts": client.session.attempts,
        "retry_used": any(a.get("retried") for a in client.session.attempts),
    }


@app.post("/api/request")
async def send_request(
    method: str = Form("GET"),
    url: str = Form(""),
    headers: str = Form("{}"),
    body: str = Form(""),
    timeout: int = Form(30),
    retry: int = Form(3),
):
    """
    单接口调试 - 发送一次 HTTP 请求并返回完整结果。
    支持方法：GET / POST / PUT / PATCH / DELETE / HEAD / OPTIONS
    重试逻辑与 common/http_client.py 保持一致。
    """
    if not url:
        raise HTTPException(400, "URL 不能为空")

    method = method.upper()
    if method not in ("GET", "POST", "PUT", "PATCH", "DELETE", "HEAD", "OPTIONS"):
        raise HTTPException(400, f"不支持的 HTTP 方法: {method}")

    # 解析 headers JSON
    try:
        req_headers = json.loads(headers)
        if not isinstance(req_headers, dict):
            req_headers = {"Content-Type": "application/json"}
    except (json.JSONDecodeError, ValueError):
        req_headers = {"Content-Type": "application/json"}

    # 限制
    timeout = max(1, min(timeout, 300))
    retry = max(0, min(retry, 10))

    result = _single_request(
        method=method,
        url=url,
        headers=req_headers,
        body=body,
        timeout=timeout,
        retry=retry,
    )
    return result


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8899)
