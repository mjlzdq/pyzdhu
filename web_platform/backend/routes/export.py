"""
结果导出路由 - /api/export
"""
import json
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Form
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import openpyxl.utils

router = APIRouter()


@router.post("/api/export")
async def api_export_results(results_json: str = Form(...)):
    """
    导出测试结果为 Excel
    包含：汇总概览 + 详细结果 + 变量池
    """
    data = json.loads(results_json)
    results = data.get("results", [])
    summary = data.get("status_counts", {})
    total = data.get("total", 0)
    pool = data.get("variable_pool", {})

    wb = openpyxl.Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    pass_fill = PatternFill(start_color="E6FFED", end_color="E6FFED", fill_type="solid")
    fail_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

    ws1 = wb.active
    ws1.title = "汇总概览"
    ws1.merge_cells('A1:B1')
    ws1['A1'] = "接口批量测试报告"
    ws1['A1'].font = Font(bold=True, size=16, color="667EEA")
    ws1['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    ws1['A4'] = "指标"
    ws1['B4'] = "数值"
    for cell in [ws1['A4'], ws1['B4']]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    metrics = [
        ("总用例数", total),
        ("成功", summary.get("成功", 0)),
        ("失败", summary.get("失败", 0)),
        ("异常", summary.get("异常", 0)),
    ]
    pass_rate = round(summary.get("成功", 0) / max(total, 1) * 100, 1)
    metrics.append(("通过率", f"{pass_rate}%"))

    for i, (label, value) in enumerate(metrics):
        ws1[f'A{5 + i}'] = label
        ws1[f'B{5 + i}'] = value
        ws1[f'A{5 + i}'].border = thin_border
        ws1[f'B{5 + i}'].border = thin_border

    if pool:
        row_offset = 5 + len(metrics) + 1
        ws1.merge_cells(f'A{row_offset}:B{row_offset}')
        ws1[f'A{row_offset}'] = "变量池"
        ws1[f'A{row_offset}'].font = Font(bold=True, size=12)
        row_offset += 1
        ws1[f'A{row_offset}'] = "变量名"
        ws1[f'B{row_offset}'] = "变量值"
        for cell in [ws1[f'A{row_offset}'], ws1[f'B{row_offset}']]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        row_offset += 1
        for k, v in pool.items():
            ws1[f'A{row_offset}'] = k
            ws1[f'B{row_offset}'] = str(v)
            ws1[f'A{row_offset}'].border = thin_border
            ws1[f'B{row_offset}'].border = thin_border
            row_offset += 1

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 40

    ws2 = wb.create_sheet("详细结果")
    detail_headers = [
        "序号", "用例ID", "描述", "方法", "URL", "状态码",
        "耗时(ms)", "结果", "失败原因", "响应体"
    ]
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for r_idx, r in enumerate(results):
        row_num = r_idx + 2
        values = [
            r_idx + 1,
            r.get("case_id", ""),
            r.get("description", ""),
            r.get("method", ""),
            r.get("url", ""),
            r.get("status_code", ""),
            r.get("elapsed_ms", ""),
            "✓ 通过" if r.get("passed") else "✗ 失败",
            r.get("failure_reason", ""),
            (r.get("response_full", "") or "")[:800],
        ]
        for col_idx, v in enumerate(values, 1):
            cell = ws2.cell(row=row_num, column=col_idx, value=v)
            cell.border = thin_border
            if col_idx == 8:
                cell.fill = pass_fill if r.get("passed") else fail_fill

    col_widths = [6, 22, 18, 8, 50, 8, 10, 10, 35, 50]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=test_report.xlsx"
        },
    )
